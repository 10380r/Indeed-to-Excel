## -*- cording: utf-8 -*-
import urllib
import urllib.request
import datetime
import time
from bs4 import BeautifulSoup
import openpyxl as excel
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import sys
import traceback
import concurrent.futures

MAX_ROWS = 220000


# 勝手に追加シリーズ: ログ管理のやつ
class Log:
    __level_dict = {
        "DEBUG": 0B100000,
        "NOTIFY": 0B10000,
        "WARNING": 0B1000,
        "ERROR": 0B100,
        "CRIT": 0B10,
        "BUG": 0B1
    }

    level_lower_all = lambda x: int("".join(['1' for _ in range(len(bin(Log.__level_dict[x])[2:]))]), 2)
    level = 0B111

    @staticmethod
    def debug(mesg):
        if Log.level & Log.__level_dict['DEBUG']:
            print(mesg)

    @staticmethod
    def notify(mesg):
        if Log.level & Log.__level_dict['NOTIFY']:
            print(mesg)

    @staticmethod
    def warning(mesg):
        if Log.level & Log.__level_dict['WARNING']:
            print(mesg)

    @staticmethod
    def error(mesg):
        if Log.level & Log.__level_dict['ERROR']:
            sys.stderr.write(str(mesg) + "\n")

    @staticmethod
    def crit(mesg):
        if Log.level & Log.__level_dict['CRIT']:
            sys.stderr.write(str(mesg) + "\n")

    @staticmethod
    def bug(mesg):
        if Log.level & Log.__level_dict['BUG']:
            sys.stderr.write(str(mesg) + "\n")


# 勝手に追加シリーズ: notify以下のレベルのログは出力しない
Log.level = Log.level_lower_all('NOTIFY')

# 新規ワークブックオブジェクトを生成する
wb = excel.Workbook()
# アクティブシートを得る
ws = wb.active
# シート名を変更する
ws.title = "渋谷区"
# B2セルにタイトルを書く。フォントサイズを24にして、センタリングする
ws['A2'] = "渋谷区の求人一覧"
ws['A2'].font = Font(size=18)
ws['A2'].alignment = Alignment(wrap_text=False,  # 折り返し改行
                               horizontal='center',  # 水平位置
                               vertical='center'  # 上下位置
                               )
# セルを結合する：　title
ws.merge_cells('A2:C2')
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 50
ws['A3'].number_format = 'yyyy-mm-dd'
ws['A3'] = datetime.date.today()
ws['A4'] = '社名'
ws['B4'] = '内容'
ws['C4'] = '場所'
ws['D4'] = '給与'
ws['E4'] = 'URL'

def scrall(url):
    tpe = concurrent.futures.ThreadPoolExecutor(max_workers=64)
    url_for_loop = url + '&limit=50&start='

    html = urllib.request.urlopen(url)
    soup = BeautifulSoup(html, 'html.parser')
    search_count = str(soup.find('div', id='searchCount'))

    result_count_unfixed = re.search(r'(求人検索結果 ([0-9,]*) 件)', search_count).group(2)
    result_count = int("".join(result_count_unfixed.split(",")))

    ## 追加 実際に取得した件数を保持
    executing_future = []
    for num in range(0, min(result_count, MAX_ROWS), 50):
        def scr_per_page(url):
            result = []
            try:
                html = urllib.request.urlopen(url)
                time.sleep(0.1)
            except:
                Log.debug("Retry %s" % (url, ))
                time.sleep(3)
                return scr_per_page(url)
            soup = BeautifulSoup(html, 'html.parser')
            div = soup.findAll('div', class_='row' or 'result' or 'clickcard')
            c = 0
            for i in div:
                if i.find('span', class_='sponsoredGray'):
                    continue

                try:
                    name = i.find('a').text
                except AttributeError:
                    name = 'undefined'

                try:
                    link = 'https://jp.indeed.com/' + i.a.get("href")
                except AttributeError:
                    link = 'undefined'

                try:
                    shop = i.find('span', class_='company').text.replace('\n', '').replace(' ', '')
                except AttributeError:
                    shop = 'undefined'

                try:
                    loc = i.find('span', class_='location').text
                except AttributeError:
                    loc = 'undefined'

                try:
                    salary = i.find('span', class_='no-wrap').text.replace('\n', '').replace(' ', '')
                except AttributeError:
                    salary = 'undefined'

                try:
                    result.append([shop, name, loc, salary, link])
                    Log.debug(separate)
                    Log.debug('社名：%s\n内容：%s\n給料：%s\n場所：%s\nURL：%s\n' % (shop, name, salary, loc, link))
                    Log.debug(url)
                    c+=1

                except AttributeError:
                    # Log.warning("AttributeError")
                    Log.debug(traceback.format_exc())
                    continue
            sys.stdout.write('.')
            sys.stdout.flush()
            return result

        executing_future.append(tpe.submit(scr_per_page, url_for_loop + str(num)))
    tpe.shutdown()
    Log.notify("取得完了")

    j = 5
    for proc in executing_future:
        for line in proc.result():
            ws['A' + str(j)] = line[0]
            ws['B' + str(j)] = line[1]
            ws['C' + str(j)] = line[2]
            ws['D' + str(j)] = line[3]
            ws['E' + str(j)] = line[4]
            j += 1
    wb.save('./results/test.xlsx')


scrall(input('URLを入力してください: '))
