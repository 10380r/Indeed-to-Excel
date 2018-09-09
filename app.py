# -*- cording: utf-8 -*-
import urllib
import datetime
from bs4 import BeautifulSoup
import openpyxl as excel
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import sys


# 新規ワークブックオブジェクトを生成する
wb = excel.Workbook()
# アクティブシートを得る
ws = wb.active
# シート名を変更する
ws.title = "sheet 1"
# A2セルにタイトルを書く。フォントサイズを24にして、センタリングする
ws['A2'] = '案件一覧'
ws['A2'].font = Font(size=18)
ws['A2'].alignment = Alignment(wrap_text=False,  # 折り返し改行
                                  horizontal='center',  # 水平位置
                                  vertical='center'  # 上下位置
                                  )
# セルを結合する
ws.merge_cells('A2:C2')
# フォーマットを整える
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 80
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws['A3'].number_format = 'yyyy-mm-dd'
ws['A3'] = datetime.date.today()

fill = excel.styles.PatternFill(patternType='solid',
                                   fgColor='FFFACD', bgColor='FFFACD')
alignment = Alignment('center')
# ws['A3'].fill = fill
ws['A5'].fill = fill
ws['B5'].fill = fill
ws['C5'].fill = fill
ws['D5'].fill = fill
ws['E5'].fill = fill
ws['A3'].alignment = Alignment(indent=3)
ws['A5'].alignment = Alignment(indent=3)
ws['B5'].alignment = Alignment(indent=3)
ws['C5'].alignment = Alignment(indent=3)
ws['D5'].alignment = Alignment(indent=3)
ws['E5'].alignment = Alignment(indent=3)
ws['A5'] = '社名'
ws['B5'] = '内容'
ws['C5'] = '場所'
ws['D5'] = '給与'
ws['E5'] = 'URL'


# スクレイピングする
def scr(url,get_page,savename):
    # URLの再利用可能にするために整える
    url_for_loop = url + '&start='
    # セルの初期値の座標を変数に格納しておく
    j = 6

    #
    for num in range(10,((get_page*10)-9), 10):
        # urllibでurlの情報を格納
        html = urllib.request.urlopen(url)
        # BeautifulSoupでスクレイピング
        soup = BeautifulSoup(html, 'html.parser')
        # 欲しい要素のタグとクラスを指定
        div = soup.findAll('div', class_='row' or 'result' or 'clickcard')
        # URLを変更
        url = url_for_loop + str(num)


        # 取得した要素から細かく要素をパースする
        # 欲しい要素がない場合がない場合にAttributeErrorを吐かれるのでエラーハンドリング
        for i in div:
            # スポンサーを除外
            if i.find('span', class_='sponsoredGray'):
                continue

            try:
                name = i.find('a').text
            except AttributeError:
                name = '-'

            try:
                link = 'https://jp.indeed.com/' + i.a.get("href")
            except AttributeError:
                link = '-'

            try:
                shop = i.find('span', class_='company').text.replace('\n', '').replace(' ', '')
            except AttributeError:
                shop = '-'

            try:
                loc = i.find('span', class_='location').text
            except AttributeError:
                loc = '-'

            try:
                salary = i.find('span', class_='no-wrap').text.replace('\n', '').replace(' ', '')
            except AttributeError:
                salary = '-'

            # エクセル側に記入していく
            # セルのrowを1づつ下にずらして記入していく
            ws['A' + str(j)] = shop
            ws['B' + str(j)] = name
            ws['C' + str(j)] = loc
            ws['D' + str(j)] = salary
            ws['E' + str(j)] = 'Click Here!'
            ws['E' + str(j)].hyperlink = link
            j += 1
        # 実行中のUI?をそれっぽく
        sys.stdout.write('.')
        sys.stdout.flush()

    # 保存処理
    wb.save('./results/'+savename+'.xlsx')
    print('Done!')

scr(input('URLを入力してください：'),int(input('何ページ分取得しますか？:')),input('ファイル名:'))
